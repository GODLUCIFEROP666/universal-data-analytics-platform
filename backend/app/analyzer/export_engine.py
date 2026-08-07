import io
from typing import Any, Dict, List

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def generate_excel_export(df: pd.DataFrame, summary_stats: Dict[str, Any] | None = None) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.copy().to_excel(writer, sheet_name="Processed Data", index=False)

        if summary_stats:
            _write_summary_sheet(writer, summary_stats)
            _write_column_sheet(writer, summary_stats.get("columns", {}))
            _write_correlation_sheet(writer, summary_stats.get("correlations", {}))
            _write_insight_sheet(writer, summary_stats.get("insights", []))

        try:
            _format_workbook(writer)
        except Exception:
            # Graceful fallback if openpyxl internals change across versions
            pass

    return output.getvalue()


def generate_csv_export(df: pd.DataFrame) -> bytes:
    """Export the current dataframe as a CSV file."""
    output = io.BytesIO()
    df.to_csv(output, index=False, encoding="utf-8-sig")
    return output.getvalue()


def _write_summary_sheet(writer: pd.ExcelWriter, summary_stats: Dict[str, Any]) -> None:
    summary = summary_stats.get("summary", {})
    rows: List[Dict[str, Any]] = [
        {"Metric": "Total Rows", "Value": summary.get("total_rows", 0)},
        {"Metric": "Total Columns", "Value": summary.get("total_columns", 0)},
        {"Metric": "Missing Values", "Value": summary.get("missing_values", 0)},
        {"Metric": "Missing Values %", "Value": summary.get("missing_percentage", 0.0)},
        {"Metric": "Duplicate Rows", "Value": summary.get("duplicate_rows", 0)},
        {"Metric": "Memory Usage (MB)", "Value": summary.get("memory_usage_mb", 0.0)},
        {"Metric": "Numeric Columns", "Value": summary.get("numeric_columns_count", 0)},
        {"Metric": "Categorical Columns", "Value": summary.get("categorical_columns_count", 0)},
        {"Metric": "Date Columns", "Value": summary.get("date_columns_count", 0)},
        {"Metric": "Analysis Time (ms)", "Value": summary.get("analysis_time_ms", 0.0)},
    ]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Executive Summary", index=False)


def _write_column_sheet(writer: pd.ExcelWriter, columns: Dict[str, Any]) -> None:
    rows: List[Dict[str, Any]] = []
    for name, info in columns.items():
        stats = info.get("stats", {})
        rows.append(
            {
                "Column": name,
                "Type": info.get("detected_type"),
                "Null Count": info.get("null_count", 0),
                "Null %": info.get("null_percentage", 0.0),
                "Unique Values": info.get("unique_count", 0),
                "Min": stats.get("min"),
                "Max": stats.get("max"),
                "Mean": stats.get("mean"),
                "Median": stats.get("median"),
                "Mode": stats.get("mode"),
                "Std Dev": stats.get("std_dev"),
                "IQR": stats.get("iqr"),
                "Outliers": stats.get("outliers_count"),
            }
        )
    if rows:
        pd.DataFrame(rows).to_excel(writer, sheet_name="Column Details", index=False)


def _write_correlation_sheet(writer: pd.ExcelWriter, correlations: Dict[str, Any]) -> None:
    if not correlations or not correlations.get("columns"):
        return
    corr_df = pd.DataFrame(correlations["values"], columns=correlations["columns"], index=correlations["columns"])
    corr_df.to_excel(writer, sheet_name="Correlation Matrix")


def _write_insight_sheet(writer: pd.ExcelWriter, insights: Any) -> None:
    rows: List[Dict[str, Any]] = []
    for insight in insights or []:
        rows.append(
            {
                "Type": insight.get("type"),
                "Severity": insight.get("severity"),
                "Title": insight.get("title"),
                "Description": insight.get("description"),
            }
        )
    if rows:
        pd.DataFrame(rows).to_excel(writer, sheet_name="Insights", index=False)


def _format_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    fill = PatternFill("solid", fgColor="0F172A")
    title_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                except Exception:
                    pass
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
